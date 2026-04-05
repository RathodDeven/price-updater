from __future__ import annotations

import subprocess
import sys
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "scripts" / "extract_price_table.py"


def _extract_target_page(input_pdf: Path, target_page: int, output_xlsx: Path) -> None:
    subprocess.run(
        [
            sys.executable,
            str(SCRIPT),
            "--input-pdf",
            str(input_pdf),
            "--output-xlsx",
            str(output_xlsx),
            "--target-page",
            str(target_page),
        ],
        cwd=ROOT,
        check=True,
    )


def _read_aliases(xlsx_path: Path) -> list[str]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    aliases: list[str] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        aliases.append(str(row[0]).strip())
    return aliases


def _read_alias_purchase(xlsx_path: Path) -> list[tuple[str, float]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[tuple[str, float]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None or row[1] is None:
            continue
        alias = str(row[0]).strip()
        try:
            purchase = float(str(row[1]).strip())
        except (TypeError, ValueError):
            continue
        rows.append((alias, purchase))
    return rows


def test_sample_1_page_7_extracts_expected_rows(tmp_path: Path) -> None:
    out_xlsx = tmp_path / "sample_1_p7.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_1.pdf", 7, out_xlsx)
    aliases = _read_aliases(out_xlsx)
    assert len(aliases) == 18
    assert "5SV53120RC" in aliases
    assert "5SV56460RC" in aliases


def test_sample_1_page_8_extracts_expected_rows(tmp_path: Path) -> None:
    out_xlsx = tmp_path / "sample_1_p8.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_1.pdf", 8, out_xlsx)
    aliases = _read_aliases(out_xlsx)
    assert len(aliases) == 30
    assert "5TJ83110" in aliases
    assert "5TJ86460" in aliases


def test_sample_1_page_9_extracts_expected_rows(tmp_path: Path) -> None:
    out_xlsx = tmp_path / "sample_1_p9.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_1.pdf", 9, out_xlsx)
    aliases = _read_aliases(out_xlsx)
    assert len(aliases) == 24
    assert "5TE3125-0RC" in aliases
    assert "5TE3492-0RC" in aliases


def test_sample_1_page_10_extracts_second_table(tmp_path: Path) -> None:
    out_xlsx = tmp_path / "sample_1_p10.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_1.pdf", 10, out_xlsx)
    aliases = _read_aliases(out_xlsx)
    assert len(aliases) == 29
    assert "8GB30101RC04" in aliases
    assert "8GB30202RC12" in aliases


def test_sample_4_page_79_includes_left_column_aliases(tmp_path: Path) -> None:
    out_xlsx = tmp_path / "sample_4_p79.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_4.pdf", 79, out_xlsx)
    aliases = _read_aliases(out_xlsx)
    # Page includes one valid left table plus right accessory table; avoid
    # counting non-priced CTX power matrix rows as aliases.
    assert len(aliases) >= 17
    assert "416870" in aliases
    assert "416873" in aliases
    assert "416874" in aliases
    assert "416889" in aliases


def test_sample_4_page_29_skips_when_purchase_column_empty(tmp_path: Path) -> None:
    out_xlsx = tmp_path / "sample_4_p29.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_4.pdf", 29, out_xlsx)
    aliases = _read_aliases(out_xlsx)
    assert aliases == []


def test_sample_4_page_31_skips_description_numeric_false_purchase(tmp_path: Path) -> None:
    out_xlsx = tmp_path / "sample_4_p31.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_4.pdf", 31, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)
    assert ("507805", 100.0) not in rows
    assert ("507806", 2762.0) in rows


def test_sample_4_page_45_skips_description_text_as_alias(tmp_path: Path) -> None:
    out_xlsx = tmp_path / "sample_4_p45.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_4.pdf", 45, out_xlsx)
    aliases = _read_aliases(out_xlsx)
    garbage = [a for a in aliases if "TERMINAL" in a.upper() or "FRONT" in a.upper()]
    assert garbage == []
    assert any(a for a in aliases if a.startswith("026") or a.startswith("042"))


def test_sample_4_page_54_uses_mrp_not_pack_values(tmp_path: Path) -> None:
    out_xlsx = tmp_path / "sample_4_p54.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_4.pdf", 54, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)
    row_set = {(alias, purchase) for alias, purchase in rows}

    # Rotary handle rows must use MRP values, not pack=1.
    assert ("668770", 1330.0) in row_set
    assert ("668772", 1350.0) in row_set
    assert ("668780", 2390.0) in row_set
    assert ("668771", 2030.0) in row_set
    assert ("668773", 2100.0) in row_set
    assert ("668781", 3940.0) in row_set

    assert ("668770", 1.0) not in row_set
    assert ("668772", 1.0) not in row_set
    assert ("668780", 1.0) not in row_set


def test_sample_4_page_46_extracts_merged_supply_inverter_row(tmp_path: Path) -> None:
    out_xlsx = tmp_path / "sample_4_p46.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_4.pdf", 46, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)
    row_set = {(alias, purchase) for alias, purchase in rows}

    # Merged text row contains alias + description + MRP in one cell.
    assert ("026410", 14070.0) in row_set
    # Row where alias is in particulars column and MRP is in mapped purchase column.
    assert ("026266", 500.0) in row_set
    # Neighbor row with missing MRP must not produce a false purchase.
    assert not any(alias == "026405" for alias, _ in rows)


def test_sample_4_page_54_extracts_phase_barrier_aliases(tmp_path: Path) -> None:
    """Phase Barriers table has Cat.No/MRP/Pack stacked in one cell."""
    out_xlsx = tmp_path / "sample_4_p54_phase.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_4.pdf", 54, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)
    row_set = {(alias, purchase) for alias, purchase in rows}

    assert ("669300", 180.0) in row_set
    assert ("669301", 230.0) in row_set
    assert ("026230", 260.0) in row_set


def test_sample_4_page_55_uses_mrp_not_pack(tmp_path: Path) -> None:
    """Stacked Cat.No/MRP/Pack cells must use MRP, not pack=1."""
    out_xlsx = tmp_path / "sample_4_p55.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_4.pdf", 55, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)
    row_set = {(alias, purchase) for alias, purchase in rows}

    # Key Locks tables must have real MRP values
    assert ("668774", 2680.0) in row_set
    assert ("668777", 2640.0) in row_set
    assert ("668782", 2640.0) in row_set
    # Must NOT have pack=1 as purchase
    assert ("668774", 1.0) not in row_set
    assert ("668777", 1.0) not in row_set
    # Padlocks table
    assert ("027180", 1810.0) in row_set


def test_sample_4_page_63_no_garbage_aliases(tmp_path: Path) -> None:
    """Complex multi-section table must not produce cross-line or description aliases."""
    out_xlsx = tmp_path / "sample_4_p63.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_4.pdf", 63, out_xlsx)
    aliases = _read_aliases(out_xlsx)

    # No garbage from cross-line alias group matching
    garbage = [a for a in aliases if len(a) > 8 and a.isdigit()]
    assert garbage == [], f"Long numeric garbage aliases: {garbage}"

    # No description text as alias
    assert "IP20" not in aliases
    assert not any("SHORTING" in a.upper() for a in aliases)

    # Valid aliases present
    assert "424006" in aliases
    assert "424108" in aliases
    assert "424201" in aliases


def test_sample_4_page_66_skips_price_on_request(tmp_path: Path) -> None:
    """Page with ■ price-on-request markers should not extract current ratings as MRP."""
    out_xlsx = tmp_path / "sample_4_p66.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_4.pdf", 66, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)

    # Should not have current ratings as purchase values
    bad_purchases = [p for _, p in rows if p in (63.0, 125.0, 4.0, 1.0)]
    assert bad_purchases == [], f"Current ratings extracted as MRP: {bad_purchases}"


def test_sample_4_page_77_uses_mrp_not_current_ratings(tmp_path: Path) -> None:
    """Thermal relays page with I min/I max columns must use MRP, not current ratings."""
    out_xlsx = tmp_path / "sample_4_p77.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_4.pdf", 77, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)
    row_dict = {alias: purchase for alias, purchase in rows}

    # Left-side "Standard type" entries: MRP = 2200
    assert row_dict.get("416640") == 2200.0
    assert row_dict.get("416645") == 2200.0
    assert row_dict.get("416649") == 2200.0

    # Right-side "Differential type": MRP = 2690-3480
    assert row_dict.get("416660") == 2690.0
    assert row_dict.get("416676") == 3480.0

    # RTX³ 65 section: MRP = 4430
    assert row_dict.get("416686") == 4430.0
    assert row_dict.get("416690") == 4430.0

    # RTX³ 100 section: MRP = 7170
    assert row_dict.get("416728") == 7170.0
    assert row_dict.get("416731") == 7170.0

    # Must NOT have current ratings as purchase
    current_ratings = {0.1, 0.16, 0.25, 0.4, 0.63, 1.0, 1.6, 2.5, 4.0, 5.0, 6.0,
                       7.0, 9.0, 12.0, 16.0, 18.0, 22.0, 28.0, 54.0, 63.0, 70.0, 80.0}
    bad = [(a, p) for a, p in rows if p in current_ratings]
    assert bad == [], f"Current ratings extracted as MRP: {bad}"


def test_sample_4_page_42_uses_mrp_not_current_ratings(tmp_path: Path) -> None:
    """MCCBs page where 'Rated Current (A)' column must NOT be used as purchase."""
    out_xlsx = tmp_path / "sample_4_p42.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_4.pdf", 42, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)
    row_dict = {alias: purchase for alias, purchase in rows}

    # Main MCCB product rows (top table) — previously missing
    assert row_dict.get("420710") == 15670.0
    assert row_dict.get("420711") == 15670.0
    assert row_dict.get("420714") == 17820.0
    assert row_dict.get("420718") == 30970.0
    assert row_dict.get("420720") == 32750.0
    assert row_dict.get("420721") == 34520.0
    assert row_dict.get("420722") == 30270.0
    assert row_dict.get("420725") == 34050.0

    # Accessory rows (accessory sub-tables)
    assert row_dict.get("420160") == 1540.0
    assert row_dict.get("420161") == 2230.0
    assert row_dict.get("421060") == 42400.0
    assert row_dict.get("422624") == 12260.0

    # Merged Motor Operator blocks with POR markers must pair price to the
    # correct trailing alias, not the first alias in the stacked sequence.
    assert row_dict.get("026144") == 62010.0
    assert row_dict.get("026126") == 71800.0
    assert row_dict.get("026123") == 71800.0
    assert row_dict.get("026127") == 72480.0
    assert "026140" not in row_dict
    assert "026124" not in row_dict
    assert "026119" not in row_dict

    # Must NOT have current ratings as purchase
    current_ratings = {16.0, 25.0, 32.0, 40.0, 50.0, 63.0, 80.0, 100.0, 125.0, 160.0}
    bad = [(a, p) for a, p in rows if p in current_ratings]
    assert bad == [], f"Current ratings extracted as MRP: {bad}"

    # Shifted dual-role motor-operator block: leading MRP in shared purchase
    # cell belongs to left alias (421060), not first in-cell alias (421061).
    assert row_dict.get("421061") == 38080.0

    # Should have 60+ rows (both main product table and all accessory sub-tables)
    assert len(rows) >= 60, f"Expected ≥60 rows, got {len(rows)}"


def test_main_price_list_page_46_skips_alias_purchase_inversions(tmp_path: Path) -> None:
    out_xlsx = tmp_path / "main_price_list_2026_p46.xlsx"
    _extract_target_page(ROOT / "samples" / "main-price-list-2026.pdf", 46, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)
    row_set = {(alias, purchase) for alias, purchase in rows}
    aliases = {alias for alias, _ in rows}

    assert ("625018", 1380.0) in row_set
    assert "DPX250ER" not in aliases
    assert "12VAC/DC" not in aliases
    assert "VOL.24VAC/DC" not in aliases
    assert "48VAC/DC" not in aliases


def test_main_price_list_page_44_skips_price_on_request_mccb_grid(tmp_path: Path) -> None:
    out_xlsx = tmp_path / "main_price_list_2026_p44.xlsx"
    _extract_target_page(ROOT / "samples" / "main-price-list-2026.pdf", 44, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)

    assert rows == []


def test_main_price_list_page_45_skips_price_on_request_mccb_grid(tmp_path: Path) -> None:
    out_xlsx = tmp_path / "main_price_list_2026_p45.xlsx"
    _extract_target_page(ROOT / "samples" / "main-price-list-2026.pdf", 45, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)

    assert rows == []


def test_main_price_list_page_37_skips_description_model_names_as_aliases(tmp_path: Path) -> None:
    out_xlsx = tmp_path / "main_price_list_2026_p37.xlsx"
    _extract_target_page(ROOT / "samples" / "main-price-list-2026.pdf", 37, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)
    row_set = {(alias, purchase) for alias, purchase in rows}
    aliases = {alias for alias, _ in rows}

    assert ("028300", 42090.0) in row_set
    assert ("028302", 52480.0) in row_set
    assert "MP2.10" not in aliases
    assert "MP4.10" not in aliases


def test_main_price_list_page_95_skips_contact_configuration_alias_garbage(tmp_path: Path) -> None:
    out_xlsx = tmp_path / "main_price_list_2026_p95.xlsx"
    _extract_target_page(ROOT / "samples" / "main-price-list-2026.pdf", 95, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)
    row_set = {(alias, purchase) for alias, purchase in rows}
    aliases = {alias for alias, _ in rows}

    assert ("417150", 610.0) in row_set
    assert ("417151", 610.0) in row_set
    assert ("417153", 940.0) in row_set
    assert ("417158", 610.0) in row_set

    for bad_alias in {"2NO", "2NC", "4NO", "4NC", "1NO", "1NC"}:
        assert bad_alias not in aliases


def test_main_price_list_page_114_uses_mrp_not_nominal_rating(tmp_path: Path) -> None:
    out_xlsx = tmp_path / "main_price_list_2026_p114.xlsx"
    _extract_target_page(ROOT / "samples" / "main-price-list-2026.pdf", 114, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)
    row_dict = dict(rows)

    assert row_dict.get("408848") == 1267.0
    assert ("408848", 60.0) not in rows


def test_main_price_list_page_138_skips_blank_mrp_rows(tmp_path: Path) -> None:
    out_xlsx = tmp_path / "main_price_list_2026_p138.xlsx"
    _extract_target_page(ROOT / "samples" / "main-price-list-2026.pdf", 138, out_xlsx)
    aliases = _read_aliases(out_xlsx)

    assert "001956" not in aliases


def test_main_price_list_pages_154_to_156_keep_split_alias_rows(tmp_path: Path) -> None:
    expected = {
        154: {("AC21104MW", 264.0), ("AC21102MW", 488.0), ("AC20109MW", 3444.0), ("AC23107MW", 648.0)},
        155: {("AC24103MW", 1996.0), ("AC24105MW", 12810.0)},
        156: {("AC23105MW", 3070.0)},
    }

    for page, expected_rows in expected.items():
        out_xlsx = tmp_path / f"main_price_list_2026_p{page}.xlsx"
        _extract_target_page(ROOT / "samples" / "main-price-list-2026.pdf", page, out_xlsx)
        rows = _read_alias_purchase(out_xlsx)
        row_set = {(alias, purchase) for alias, purchase in rows}
        for expected_row in expected_rows:
            assert expected_row in row_set


def test_main_price_list_page_156_recovers_both_spread_blocks(tmp_path: Path) -> None:
    out_xlsx = tmp_path / "main_price_list_2026_p156_spread.xlsx"
    _extract_target_page(ROOT / "samples" / "main-price-list-2026.pdf", 156, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)
    row_set = {(alias, purchase) for alias, purchase in rows}
    aliases = {alias for alias, _ in rows}

    expected_rows = {
        ("AC27108MW", 3298.0),
        ("AC27111MW", 9388.0),
        ("679255", 234.0),
        ("076711", 2698.0),
        ("AC361MW", 9354.0),
        ("AC20110MW", 11044.0),
        ("AC22114MW", 812.0),
        ("AC23115MW", 794.0),
        ("AC21117MW", 778.0),
        ("AC23119MW", 778.0),
        ("AC657MW", 4188.0),
        ("AC627MW", 6692.0),
        ("AC20111MW", 1332.0),
        ("AC20112MW", 1368.0),
        ("AC628MW", 796.0),
    }

    for expected_row in expected_rows:
        assert expected_row in row_set

    assert "2MODULE" not in aliases
    assert len(rows) >= 28


def test_main_price_list_page_157_recovers_merged_right_block(tmp_path: Path) -> None:
    out_xlsx = tmp_path / "main_price_list_2026_p157_spread.xlsx"
    _extract_target_page(ROOT / "samples" / "main-price-list-2026.pdf", 157, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)
    row_set = {(alias, purchase) for alias, purchase in rows}
    aliases = {alias for alias, _ in rows}

    expected_rows = {
        ("AC625MW", 796.0),
        ("AC626MW", 796.0),
        ("AC4430MW", 6684.0),
        ("AC4421MW", 9306.0),
        ("AC4411MW", 11044.0),
        ("AC4400MW", 1180.0),
        ("AC4452MW", 9484.0),
        ("AC350MW", 1810.0),
        ("AC354MW", 576.0),
        ("AC150MW", 82.0),
        ("AC153MW", 122.0),
        ("A1401", 500.0),
        ("AC21104MB", 306.0),
        ("AC21106MB", 558.0),
        ("AC22109MB", 1036.0),
    }

    for expected_row in expected_rows:
        assert expected_row in row_set

    assert "2MODULE" not in aliases
    assert len(rows) >= 33


def test_sample_4_page_53_avoids_current_leak_for_alias_669198(tmp_path: Path) -> None:
    """When alias has competing candidates, prefer MRP over rated-current leakage."""
    out_xlsx = tmp_path / "sample_4_p53.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_4.pdf", 53, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)
    row_dict = {alias: purchase for alias, purchase in rows}

    # This alias appears in a row where current(250) and MRP(54260) coexist.
    # Extraction must keep the MRP pair, not the current value.
    assert row_dict.get("669198") == 54260.0

    # Surrounding rows in same block stay correct.
    assert row_dict.get("669197") == 48230.0
    assert row_dict.get("669207") == 57270.0
    assert row_dict.get("669208") == 64010.0


def test_sample_4_page_78_avoids_current_leak_for_thermal_relays(tmp_path: Path) -> None:
    """Mixed Imin/Imax/MRP stacked cells must map purchase to MRP, not current."""
    out_xlsx = tmp_path / "sample_4_p78.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_4.pdf", 78, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)
    row_dict = {alias: purchase for alias, purchase in rows}

    # Thermal overload relay rows where purchase cell includes Imin/Imax/MRP/pack.
    assert row_dict.get("416780") == 14340.0
    assert row_dict.get("416781") == 14340.0
    assert row_dict.get("416782") == 15900.0
    assert row_dict.get("416783") == 15900.0
    assert row_dict.get("416784") == 16510.0
    assert row_dict.get("416786") == 21050.0
    assert row_dict.get("416787") == 21050.0
    assert row_dict.get("416788") == 21050.0
    assert row_dict.get("416789") == 21050.0


def test_sample_4_page_79_skips_ctx_kvar_matrix_garbage(tmp_path: Path) -> None:
    """CTX kVAr/current matrix has no Cat.Nos+MRP pairs and must be skipped."""
    out_xlsx = tmp_path / "sample_4_p79.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_4.pdf", 79, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)
    row_dict = {alias: purchase for alias, purchase in rows}

    # Valid rows from priced sections remain.
    assert row_dict.get("416870") == 2540.0
    assert row_dict.get("416871") == 2540.0
    assert row_dict.get("416872") == 3480.0
    assert row_dict.get("416873") == 3320.0
    assert row_dict.get("416874") == 2410.0
    assert row_dict.get("416875") == 3280.0
    assert row_dict.get("416876") == 3550.0
    assert row_dict.get("416877") == 3550.0

    # Garbage aliases from CTX matrix must not be extracted.
    for bad_alias in ["CTX322", "CTX340", "CTX365", "CTX3100", "CTX3225", "CTX3400", "CTX3800", "IEC60947-4-1"]:
        assert bad_alias not in row_dict


def test_sample_4_page_84_prefers_mrp_over_description_numbers(tmp_path: Path) -> None:
    """Split rows with alias in description must use mapped MRP, not description numerics."""
    out_xlsx = tmp_path / "sample_4_p84.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_4.pdf", 84, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)
    row_dict = {alias: purchase for alias, purchase in rows}

    assert row_dict.get("417406") == 1700.0
    assert row_dict.get("417408") == 1700.0
    assert row_dict.get("417407") == 1700.0

    # Ensure description numeric token ("63") is not used as purchase.
    assert row_dict.get("417408") != 63.0


def test_sample_4_page_94_prefers_mrp_column_over_nominal_rating(tmp_path: Path) -> None:
    """Right-side rows with merged nominal+module+alias must use mapped MRP column."""
    out_xlsx = tmp_path / "sample_4_p94.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_4.pdf", 94, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)
    row_dict = {alias: purchase for alias, purchase in rows}

    # Previously leaked nominal rating values (63/100/125) as purchase.
    assert row_dict.get("406502") == 986.0
    assert row_dict.get("406504") == 1354.0
    assert row_dict.get("406505") == 1540.0
    assert row_dict.get("406511") == 1560.0
    assert row_dict.get("406514") == 1986.0
    assert row_dict.get("408831") == 2172.0

    # Section headings must not be emitted as synthetic aliases.
    assert "DOUBLEPOLE240V" not in row_dict


def test_sample_4_page_95_recovers_split_rcb_prices(tmp_path: Path) -> None:
    """Split rows with alias and MRP separated by image-label noise should recover true MRP."""
    out_xlsx = tmp_path / "sample_4_p95.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_4.pdf", 95, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)
    row_dict = {alias: purchase for alias, purchase in rows}

    assert row_dict.get("411851") == 4022.0
    assert row_dict.get("411852") == 4670.0
    assert row_dict.get("411873") == 9146.0
    assert row_dict.get("411898") == 10148.0


def test_sample_4_page_93_skips_rows_with_por_in_mrp_column(tmp_path: Path) -> None:
    """When mapped MRP cell is POR marker, row must be skipped (no nominal fallback)."""
    out_xlsx = tmp_path / "sample_4_p93.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_4.pdf", 93, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)
    row_dict = {alias: purchase for alias, purchase in rows}

    # These aliases have POR marker in mapped MRP column on page 93.
    # Do not emit fallback purchases from nominal/current values.
    for bad_alias in ["408721", "408767", "408793", "408794"]:
        assert bad_alias not in row_dict


def test_sample_4_page_100_strips_flattened_alias_footnote_digits(tmp_path: Path) -> None:
    """Cat.Nos with superscript footnotes must not append marker digits to alias."""
    out_xlsx = tmp_path / "sample_4_p100.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_4.pdf", 100, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)
    row_set = {(alias, purchase) for alias, purchase in rows}

    # Right-side SPD table aliases carrying superscript 1 in source PDF should
    # normalize to their base Cat.Nos values.
    assert ("412276", 17732.0) in row_set
    assert ("412277", 32328.0) in row_set
    assert ("412256", 15312.0) in row_set
    assert ("412257", 27854.0) in row_set
    assert ("412281", 28980.0) in row_set
    assert ("412283", 79524.0) in row_set

    # Left-side block aliases should also be extracted.
    assert ("414446", 24738.0) in row_set
    assert ("414447", 24738.0) in row_set
    assert ("414448", 24738.0) in row_set
    assert ("414449", 24738.0) in row_set
    assert ("414261", 10408.0) in row_set
    assert ("414262", 11756.0) in row_set
    assert ("414263", 11756.0) in row_set
    assert ("414281", 20154.0) in row_set
    assert ("414282", 20110.0) in row_set
    assert ("414283", 20902.0) in row_set

    for bad_alias in ["4122761", "4122771", "4122561", "4122571", "4122811", "4122831"]:
        assert not any(alias == bad_alias for alias, _ in rows)


def test_sample_4_page_112_blanking_plate_uses_mrp_not_pack(tmp_path: Path) -> None:
    """Shifted subsection rows must not use pack as purchase (alias 601470)."""
    out_xlsx = tmp_path / "sample_4_p112.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_4.pdf", 112, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)
    row_set = {(alias, purchase) for alias, purchase in rows}

    assert ("601470", 18.0) in row_set
    assert ("601470", 10.0) not in row_set


def test_sample_4_page_129_extracts_shifted_alias_573451(tmp_path: Path) -> None:
    """Shifted rows with stale alias column must keep leading particulars Cat.No."""
    out_xlsx = tmp_path / "sample_4_p129.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_4.pdf", 129, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)
    row_set = {(alias, purchase) for alias, purchase in rows}

    assert ("573451", 480.0) in row_set
    assert ("573450", 406.0) in row_set


def test_sample_4_page_290_extracts_alias_057299_with_trailing_particulars_price(tmp_path: Path) -> None:
    """Rows with blank purchase column but trailing MRP in particulars should be recovered."""
    out_xlsx = tmp_path / "sample_4_p290.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_4.pdf", 290, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)
    row_set = {(alias, purchase) for alias, purchase in rows}

    assert ("057299", 32970.0) in row_set
    assert ("058918", 29260.0) in row_set


def test_sample_4_page_289_recovers_mrp_when_purchase_cell_has_pack_token(tmp_path: Path) -> None:
    """Shifted rows with purchase=pack token should recover MRP from particulars text."""
    out_xlsx = tmp_path / "sample_4_p289.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_4.pdf", 289, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)
    row_set = {(alias, purchase) for alias, purchase in rows}

    assert ("057720", 4170.0) in row_set
    assert ("057720", 1.0) not in row_set


def test_sample_4_page_256_recovers_mrp_for_011110_in_split_rows(tmp_path: Path) -> None:
    """Split rows should not inherit description numeric tokens as purchase."""
    out_xlsx = tmp_path / "sample_4_p256.xlsx"
    _extract_target_page(ROOT / "samples" / "sample_4.pdf", 256, out_xlsx)
    rows = _read_alias_purchase(out_xlsx)
    row_set = {(alias, purchase) for alias, purchase in rows}

    assert ("011110", 1450.0) in row_set
    assert ("011110", 65.0) not in row_set
    assert ("011162", 470.0) in row_set
    assert ("011162", 65.0) not in row_set
    assert ("638001", 210.0) in row_set
    assert ("638001", 75.0) not in row_set
    assert ("638038", 1270.0) in row_set
    assert ("638008", 240.0) in row_set
