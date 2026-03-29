from __future__ import annotations

import subprocess
import sys
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "scripts" / "extract_price_table.py"
TRUTH_DIR = ROOT / "tests" / "truth"


def _extract_to_xlsx(input_pdf: Path, output_xlsx: Path) -> None:
    subprocess.run(
        [
            sys.executable,
            str(SCRIPT),
            "--input-pdf",
            str(input_pdf),
            "--output-xlsx",
            str(output_xlsx),
        ],
        cwd=ROOT,
        check=True,
    )


def _read_rows(xlsx_path: Path) -> list[tuple[str, float, int]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[tuple[str, float, int]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        alias = str(row[0]).strip()

        purchase_cell = row[1]
        if not isinstance(purchase_cell, (int, float, str)):
            continue
        purchase = float(purchase_cell)

        source_page_cell = row[2]
        if not isinstance(source_page_cell, (int, float, str)):
            continue
        source_page = int(source_page_cell)

        rows.append((alias, purchase, source_page))
    return rows


def _assert_workbook_matches_truth(tmp_path: Path, sample_name: str) -> None:
    input_pdf = ROOT / "samples" / f"{sample_name}.pdf"
    truth_xlsx = TRUTH_DIR / f"{sample_name}_truth.xlsx"
    out_xlsx = tmp_path / f"{sample_name}_generated.xlsx"

    assert input_pdf.exists(), f"Missing sample PDF: {input_pdf}"
    assert truth_xlsx.exists(), f"Missing locked truth workbook: {truth_xlsx}"

    _extract_to_xlsx(input_pdf, out_xlsx)

    expected_rows = _read_rows(truth_xlsx)
    actual_rows = _read_rows(out_xlsx)
    assert actual_rows == expected_rows


def test_sample_2_full_workbook_regression(tmp_path: Path) -> None:
    _assert_workbook_matches_truth(tmp_path, "sample_2")


def test_sample_3_full_workbook_regression(tmp_path: Path) -> None:
    _assert_workbook_matches_truth(tmp_path, "sample_3")


def test_sample_1_full_workbook_regression(tmp_path: Path) -> None:
    _assert_workbook_matches_truth(tmp_path, "sample_1")
